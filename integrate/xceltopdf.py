import os
from flask import Blueprint, redirect, render_template, request, send_file, current_app, url_for
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from io import BytesIO

from email_utils import send_email

exceltopdf_bp = Blueprint('exceltopdf', __name__)

@exceltopdf_bp.route('/xceltopdf', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            return render_template('xceltopdf.html', error='No file part')

        excel_file = request.files['excel_file']
        
        if excel_file.filename == '':
            return render_template('xceltopdf.html', error='No selected file')

        if excel_file:
            upload_folder = os.path.join(current_app.root_path, 'uploads')
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)

            excel_path = os.path.join(upload_folder, excel_file.filename)
            excel_file.save(excel_path)
            
            pdf_filename = 'output.pdf'
            pdf_path = os.path.join(upload_folder, pdf_filename)
            convert_to_pdf(excel_path, pdf_path)
            action = request.form.get('action')
            if action == 'send':
                recipient_email = request.form.get('email')
                smtp_username = 'dconvertz@gmail.com'  
                smtp_password = 'aicwueerhuresupz' 
                send_email(pdf_path, recipient_email, smtp_username, smtp_password)
                return redirect(url_for('exceltopdf.index'))
            elif action == 'download':
                return send_file(pdf_path, as_attachment=True)
    
    return render_template('xceltopdf.html', error=None)

def convert_to_pdf(input_path, output_path):
    wb = load_workbook(input_path)
    sheet = wb.active

    data = []
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(str(cell.value))
        data.append(row_data)

    doc = SimpleDocTemplate(output_path, pagesize=letter)
    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), '#f8f8f8'),
        ('TEXTCOLOR', (0, 0), (-1, -1), (0, 0, 0)),  # black text color for the entire table
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, '#333333')
    ])
    table.setStyle(style)

    elements = [table]
    doc.build(elements)
