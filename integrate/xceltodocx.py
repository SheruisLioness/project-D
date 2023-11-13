import os
from flask import Blueprint, redirect, render_template, request, send_file, current_app, url_for, flash
from openpyxl import load_workbook
from docx import Document
from io import BytesIO

from email_utils import send_email

exceltodocx_bp = Blueprint('exceltodocx', __name__)

@exceltodocx_bp.route('/xceltodocx', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            return render_template('xceltodocx.html', error='No file part')

        excel_file = request.files['excel_file']
        
        if excel_file.filename == '':
            return render_template('xceltodocx.html', error='No selected file')

        if excel_file.filename.split('.')[-1] not in ['xls', 'xlsx']:
            flash('Only XLS and XLSX formats are supported', 'danger')
            return render_template('xceltodocx.html')

        if excel_file:
            upload_folder = os.path.join(current_app.root_path, 'uploads')
            if not os.path.exists(upload_folder):
                os.makedirs(upload_folder)

            excel_path = os.path.join(upload_folder, excel_file.filename)
            excel_file.save(excel_path)
            
            docx_filename = 'output.docx'
            docx_path = os.path.join(upload_folder, docx_filename)
            convert_to_docx(excel_path, docx_path)
            action = request.form.get('action')
            
            if action == 'send':
                recipient_email = request.form.get('email')
                smtp_username = 'dconvertz@gmail.com'  
                smtp_password = 'aicwueerhuresupz'  
                send_email(docx_path, recipient_email, smtp_username, smtp_password)
                return redirect(url_for('exceltodocx.index'))
            elif action == 'download':
                return send_file(docx_path, as_attachment=True)
    
    return render_template('xceltodocx.html', error=None)

def convert_to_docx(input_path, output_path):
    wb = load_workbook(input_path)
    sheet = wb.active

    doc = Document()
    
    # Create a table with the same number of columns as in the Excel sheet
    table = doc.add_table(rows=1, cols=len(sheet[1]))

    # Set the style of the table
    table.style = 'Table Grid'

    # Add column names as the first row in the table
    heading_cells = table.rows[0].cells
    for col_num, cell in enumerate(sheet[1], start=1):
        heading_cells[col_num-1].text = str(cell.value)

    # Add data rows to the table
    for row in sheet.iter_rows(min_row=2):
        row_cells = table.add_row().cells
        for col_num, cell in enumerate(row, start=1):
            row_cells[col_num-1].text = str(cell.value)

    doc.save(output_path)
