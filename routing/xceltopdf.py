import os
from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from io import BytesIO

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

def convert_to_pdf(input_path, output_path):
    wb = load_workbook(input_path)
    sheet = wb.active

    data = []
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(str(cell.value))
        data.append(row_data)

    doc = SimpleDocTemplate(output_path, pagesize=letter)
    table = Table(data)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), ('#333333')),
        ('TEXTCOLOR', (0, 0), (-1, 0), (1, 1, 1)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), '#f8f8f8'),
        ('GRID', (0, 0), (-1, -1), 1, '#333333')
    ])
    table.setStyle(style)

    elements = [table]
    doc.build(elements)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            return render_template('xceltopdf.html', error='No file part')

        excel_file = request.files['excel_file']
        
        if excel_file.filename == '':
            return render_template('xceltopdf.html', error='No selected file')

        if excel_file:
            excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_file.filename)
            excel_file.save(excel_path)
            
            pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], 'output.pdf')
            convert_to_pdf(excel_path, pdf_path)
            
            return render_template('download_pdf.html', pdf_path='output.pdf')
    
    return render_template('xceltopdf.html', error=None)

@app.route('/download_pdf')
def download_pdf():
    pdf_path = os.path.join(app.config['UPLOAD_FOLDER'], 'output.pdf')
    return send_file(pdf_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
