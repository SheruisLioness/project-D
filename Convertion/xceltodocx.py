import os
from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
from docx import Document

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'

def convert_to_docx(input_path, output_path):
    wb = load_workbook(input_path)
    sheet = wb.active

    doc = Document()
    
    # Create a table with the same number of columns as in the Excel sheet
    table = doc.add_table(rows=1, cols=len(sheet[1]))

    # Set the style of the table
    table.style = 'Table Grid'

    # Add column names as the first row in the table
    heading_cells = table.rows[0].cells
    for col_num, cell in enumerate(sheet[1], start=1):
        heading_cells[col_num-1].text = str(cell.value)

    # Add data rows to the table
    for row in sheet.iter_rows(min_row=2):
        row_cells = table.add_row().cells
        for col_num, cell in enumerate(row, start=1):
            row_cells[col_num-1].text = str(cell.value)

    doc.save(output_path)

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            return render_template('xceltodocx.html', error='No file part')

        excel_file = request.files['excel_file']
        
        if excel_file.filename == '':
            return render_template('xceltodocx.html', error='No selected file')

        if excel_file:
            excel_path = os.path.join(app.config['UPLOAD_FOLDER'], excel_file.filename)
            excel_file.save(excel_path)
            
            docx_path = os.path.join(app.config['UPLOAD_FOLDER'], 'output.docx')
            convert_to_docx(excel_path, docx_path)
            
            return render_template('download_docx.html', docx_path='output.docx')
    
    return render_template('xceltodocx.html', error=None)

@app.route('/download_docx')
def download_docx():
    docx_path = os.path.join(app.config['UPLOAD_FOLDER'], 'output.docx')
    return send_file(docx_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
